'''
This module is used to create excel headword files. 
Because the quality of the existing html is low and I am going to have to allow for the errors
when generating the new html

The excel files will be used to record the errors as a basis to generate high quality html

An example of a key section of the html is as follows...

<div class="section" id="Ekieki" xml:lang="en">
  <p class="hang">
    <span class="foreign bold" xml:lang="mi">Ekieki</span> =
    <b>ikeike</b>, a.
    <i>High, lofty</i>.
  </p>
</div>

The key element being the id of the div, that is "Ekieki"

'''

import config
import pprint
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import NamedStyle, Alignment, Font

PREFACE_PAGES_COUNT = 25 # the html treats page 1 as page 26, due to the preface pages xxv (25)

dictionary_letters = ('A', 'E', 'H', 'I', 'K', 'M', 'N', 'Ng', 'O', 'P', 'R', 'T', 'U', 'W', 'Wh')
start_pages = {
    'A' : 1,
    'E' : 25,
    'H' : 29,
    'I' : 73,
    'K' : 81,
    'M' : 161,
    'N' : 216,
    'Ng' : 225,
    'O' : 237,
    'P' : 243,
    'R' : 319,
    'T' : 354,
    'U' : 464,
    'W' : 472,
    'Wh' : 484,
}

colours = {
    'A' : "00247d", #Aumoana
    'E' : "a06d38",
    'H' : "c0c0c0",
    'I' : "e8e8e8",
    'K' : "ffa500",
    'M' : "ffc0cb",
    'N' : "000000",
    'Ng' : "ff0000",
    'O' : "a9a9a9",
    'P' : "ffdf00",
    'R' : "c64e59",
    'T' : "ffffff",
    'U' : "006400",
    'W' : "800080",
    'Wh' : "ff4d4d",
}


def create_excel_headword_file(letter):

    cf = config.ConfigFile()
    text_files_path = (cf.configfile[cf.computername]['original_files_path'])
    # sections_and_pbs = []

    # open the specific html file
    file_name = letter + ".html"
    text_file_path = text_files_path + file_name
    with open(text_file_path, 'r') as f:
        soup = BeautifulSoup(f)
        sections_and_pbs = soup.select(".section,[title='page break']") #pb = page break

    # get page numbers and headwords
    page_numbers_and_headwords = get_page_numbers_and_headwords(letter, sections_and_pbs)
    # od = OrderedDict(sorted(Counter(page_numbers_and_headwords).items()))
    pprint.pprint(page_numbers_and_headwords)

    # write the Excel file
    wb = Workbook()
    excel_files_path = (cf.configfile[cf.computername]['excel_files_path'])
    excel_file_name = letter + ".xlsx"
    excel_file_path = excel_files_path + excel_file_name

    page_numbers = list(set(x[0] for x in page_numbers_and_headwords))
    
    for page_counter, page_number in enumerate(page_numbers, 0):
        headwords_for_page = [x[1] for x in page_numbers_and_headwords if x[0] == page_number]

        #get worksheet title
        if page_counter == 0:
            active_worksheet = wb.active
            active_worksheet.title = str(start_pages[letter])
        else:
            active_worksheet = wb.create_sheet(title=str(start_pages[letter] + page_counter))

        # data validation
        dv = DataValidation(type="list", formula1='"yes,no,adjust"', allow_blank=True)
        active_worksheet.add_data_validation(dv)

        #named styles
        style_title = NamedStyle(name="style_title")
        style_title.font = Font(bold=True, color=colours[letter], italic=True)

        # had to wrap this in try / except because sometimes getting that style_title already existed
        # when creating the workbook!
        try:
            wb.add_named_style(style_title)
        except ValueError:
            pass

        # titles
        active_worksheet["A1"] = "Entry"
        active_worksheet["A1"].style = 'style_title'
        active_worksheet["B1"] = "Headword"
        active_worksheet["B1"].style = 'style_title'
        active_worksheet["C1"] = "Status"
        active_worksheet["C1"].style = 'style_title'
        active_worksheet["D1"] = "Adjusted"
        active_worksheet["D1"].style = 'style_title'

        for counter, headword in enumerate(headwords_for_page, 1):
            active_worksheet.cell(row=counter + 1, column=1, value=counter)
            active_worksheet.cell(row=counter + 1, column=1).font = Font(bold=True)
            active_worksheet.cell(row=counter + 1, column=2, value=headword)
            dv.add("C"+str(counter + 1)) # couldn't use row column syntax

        rows = range(1, len(headwords_for_page) + 1 + 1)
        columns = range(1, 4 + 1)
        for row in rows:
            for col in columns:
                active_worksheet.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')

        active_worksheet.sheet_format.defaultRowHeight = 22
        active_worksheet.sheet_format.baseColWidth = 12
        active_worksheet.sheet_view.zoomScale = 140
        active_worksheet.sheet_view.showGridLines = False
        active_worksheet.sheet_properties.tabColor = colours[letter]
    wb.save(filename = excel_file_path)
    return True


def get_page_numbers_and_headwords(letter, sections_and_pbs):
    '''
    Given the soup passed returns the page number of each section and the associated headword
    where the headword has 'vowel.' in it, this is changed to the 'macronised vowel'
    This is achieved by .replace(".", chr(772)). I was surprised this worked as I was just playing
    around in the editor but it appears to.
    '''   

    page_numbers_and_headwords = []

    for counter, section_or_pb in enumerate(sections_and_pbs):
        # print(counter, section_or_pb.name)
        # only going to create page numbers for sections
        # find the first page break entry (if it exists), looking backwards from where we are
        if section_or_pb.name == "div":
            reversed_list_before = reversed(sections_and_pbs[:counter])
            try:
                page_break_entry = next(x for x in reversed_list_before if x.name == "a")
                # print(page_break_entry)
            except StopIteration:
                # we are on the first page for the letter and it is not a brand new page
                page_number_to_use = start_pages[letter]
            else:
                page_number = int(page_break_entry["href"].replace('#n', ''))
                page_number_to_use = page_number - PREFACE_PAGES_COUNT

            page_numbers_and_headwords. \
                append((page_number_to_use, section_or_pb["id"].replace(".", chr(772))))

    return page_numbers_and_headwords




if __name__ == '__main__':

    import sys
    import argparse
    import ast

    # create the top-level parser
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers()

    # create the parser for the create_excel_headword_file function
    create_excel_headword_file_parser = subparsers.add_parser('create_excel_headword_file')
    create_excel_headword_file_parser.add_argument('letter', choices = list(dictionary_letters))
    create_excel_headword_file_parser.set_defaults(function = create_excel_headword_file)

    # parse the arguments
    arguments = parser.parse_args()
    arguments = vars(arguments) #convert from Namespace to dict

    #attempt to extract and then remove the function entry
    try:
        function_to_call = arguments['function']
    except KeyError:
        print ("You need a function name. Please type -h to get help")
        sys.exit()
    else:
        #remove the function entry as we are only passing arguments
        del arguments['function']

    if arguments:
        #remove any entries that have a value of 'None'
        #We are *assuming* that these are optional
        #We are doing this because we want the function definition to define
        #the defaults (NOT the function call)
        arguments = { k : v for k,v in arguments.items() if v is not None }

        #alter any string 'True' or 'False' to bools
        arguments = { k : ast.literal_eval(v) if v in ['True','False'] else v
                                              for k,v in arguments.items() }

    result = function_to_call(**arguments) #note **arguments works fine for empty dict {}

    print (result)
